"""Generate a per-file Excel report for PDF/RTF/Word/HTML QA."""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEV_COLORS={"Critical":("E53935","FFF0F0"),"Major":("B45309","FFFBF0"),"Minor":("2563EB","F0F7FF"),"Passed":("15803D","EDFFF4"),"error":("E53935","FFF0F0"),"warning":("B45309","FFFBF0"),"info":("2563EB","F0F7FF")}
BORDER=Border(*[Side(border_style="thin",color="D0D5EE")]*4)

def generate_report(source_path,target_path,issues,language,unique_id,output_path,validation_mode="pdf"):
    labels={"pdf":"PDF ↔ HTML","rtf":"RTF ↔ HTML","pdf_word":"PDF ↔ WORD"}
    mode=labels.get(validation_mode,str(validation_mode).upper())
    source_name=os.path.basename(source_path); target_name=os.path.basename(target_path)
    wb=Workbook(); ws=wb.active; ws.title="QA Report"; ws.sheet_view.showGridLines=False
    cols=[("Unique ID",14),("Validation Type",18),("Source File",34),("Target File",40),("Language",14),("Error Type",28),("Severity",12),("Target Location",16),("Description",54),("Expected",38),("Actual / Shown",38),("Excerpt / Snippet",56)]
    for i,(h,w) in enumerate(cols,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols)); c=ws.cell(1,1,f"Document QA Report · {source_name} · Generated {datetime.now():%Y-%m-%d %H:%M}"); c.font=Font(name="Arial",bold=True,color="FFFFFF",size=12); c.fill=PatternFill("solid",start_color="1E2240"); c.alignment=Alignment(vertical="center"); ws.row_dimensions[1].height=30
    total=len(issues); counts={s:sum(1 for i in issues if str(getattr(i,"severity","")).strip()==s) for s in ("Critical","Major","Minor")}
    meta=[f"ID: {unique_id}",f"Validation: {mode}",f"Source: {source_name}",f"Target: {target_name}",f"Language: {language}",f"Total Issues: {total}"]
    for i,v in enumerate(meta,1):
        c=ws.cell(2,i,v); c.font=Font(name="Arial",size=9,italic=True,color="5B6BA0"); c.fill=PatternFill("solid",start_color="EEF1FF"); c.border=BORDER
    for i,(h,_) in enumerate(cols,1):
        c=ws.cell(3,i,h); c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.fill=PatternFill("solid",start_color="1E2240"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORDER
    if not issues:
        ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=len(cols)); c=ws.cell(4,1,"✓ No issues found — content matches the compared documents."); c.font=Font(name="Arial",bold=True,color="1B7F3A"); c.fill=PatternFill("solid",start_color="EDFFF4"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORDER
    else:
        order={"Critical":0,"Major":1,"Minor":2,"error":0,"warning":1,"info":2}
        for r,issue in enumerate(sorted(issues,key=lambda x:(order.get(str(getattr(x,"severity","") or "Minor"),3),getattr(x,"line",None) or 99999)),4):
            sev=str(getattr(issue,"severity","") or "Minor"); fg,bg=SEV_COLORS.get(sev,("000000","FFFFFF"))
            vals=[unique_id,mode,source_name,target_name,language,getattr(issue,"category","") or "",sev,getattr(issue,"line",None) or "—",getattr(issue,"message","") or "",getattr(issue,"expected","") or "",getattr(issue,"actual","") or "",getattr(issue,"snippet","") or ""]
            for i,v in enumerate(vals,1):
                c=ws.cell(r,i,v); c.font=Font(name="Arial",bold=(i==7),color=(fg if i==7 else "000000")); c.fill=PatternFill("solid",start_color=bg); c.alignment=Alignment(horizontal="center" if i in (1,2,5,7,8) else "left",vertical="top",wrap_text=True); c.border=BORDER
            ws.row_dimensions[r].height=60
    ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:{get_column_letter(len(cols))}3"; wb.save(output_path); return output_path