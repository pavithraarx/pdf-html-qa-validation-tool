"""
server.py — Flask backend for Source→HTML QA Tool

Adds:
- SQLite users/login/history
- password hashing with Werkzeug
- user-owned runs and downloads
- secure upload validation
- batch_runs/ storage only
- 3-sheet Excel reports for batch and individual downloads

Core QA logic is not changed: run_checks(source_path, html_path, progress_cb=...) is still used.
"""

import io
import json
import os
import re
import sys
import tempfile
import threading
import unicodedata
import uuid
import webbrowser
import zipfile
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from types import SimpleNamespace

from flask import Flask, jsonify, request, send_file, send_from_directory, session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash

sys.path.insert(0, os.path.dirname(__file__))
from qa_engine import run_checks
import database as db

BASE = Path(__file__).resolve().parent
STATIC = BASE / "static"
BATCH_RUNS = BASE / "batch_runs"

# Do NOT auto-create old top-level reports/ or audits/ folders.
# All new generated outputs go inside batch_runs/<run>/reports|audits|metadata.
for _d in (STATIC, BATCH_RUNS):
    _d.mkdir(exist_ok=True)

VALIDATOR_VERSION = "4.1.1-single-fixed-admin-clean-auth"

app = Flask(__name__, static_folder=str(STATIC))

# Session security
# - If QA_TOOL_SECRET_KEY is not set, a fresh key is generated on every server start.
#   That forces users to log in again after the server restarts.
# - For production, set QA_TOOL_SECRET_KEY to a long random value so sessions survive restarts.
app.config["SECRET_KEY"] = os.environ.get("QA_TOOL_SECRET_KEY") or os.urandom(32)
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("QA_TOOL_MAX_UPLOAD_MB", "250")) * 1024 * 1024
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=int(os.environ.get("QA_TOOL_SESSION_MINUTES", "30")))
app.config["SESSION_REFRESH_EACH_REQUEST"] = True
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("QA_TOOL_COOKIE_SECURE", "0") == "1"

SESSION_IDLE_TIMEOUT_SECONDS = int(os.environ.get("QA_TOOL_IDLE_TIMEOUT_SECONDS", "1800"))

# Upload safety limits.
MAX_ZIP_FILES = int(os.environ.get("QA_TOOL_MAX_ZIP_FILES", "3000"))
MAX_ZIP_UNCOMPRESSED_BYTES = int(os.environ.get("QA_TOOL_MAX_ZIP_MB", "500")) * 1024 * 1024
ALLOWED_SOURCE_TYPES = {"pdf", "rtf", "pdf_word"}
ALLOWED_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}
ALLOWED_UPLOAD_EXTS = {".zip", ".pdf", ".rtf", ".docx", ".html", ".htm"} | ALLOWED_IMAGE_EXTS


# Fixed single admin account.
# The portal supports exactly ONE admin. Admin credentials are read from
# admin_config.json, with optional environment-variable overrides.
# SQLite stores only the password hash, never the plain password.
ADMIN_CONFIG_PATH = BASE / "admin_config.json"

def _load_admin_config() -> dict:
    default_cfg = {
        "email": "admin@company.local",
        "username": "admin",
        "password": "Admin@12345",
        "display_name": "System Admin",
    }
    cfg = dict(default_cfg)
    if ADMIN_CONFIG_PATH.exists():
        try:
            loaded = json.loads(ADMIN_CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                cfg.update({k: v for k, v in loaded.items() if v is not None})
        except Exception:
            # Keep defaults if config is unreadable so the server can still start.
            pass
    else:
        try:
            ADMIN_CONFIG_PATH.write_text(json.dumps(default_cfg, indent=2), encoding="utf-8")
        except Exception:
            pass

    # Optional overrides for local testing/deployment.
    cfg["email"] = os.environ.get("QA_ADMIN_EMAIL", cfg["email"])
    cfg["username"] = os.environ.get("QA_ADMIN_USERNAME", cfg["username"])
    cfg["password"] = os.environ.get("QA_ADMIN_PASSWORD", cfg["password"])
    cfg["display_name"] = os.environ.get("QA_ADMIN_DISPLAY_NAME", cfg.get("display_name") or "System Admin")
    return cfg

_ADMIN_CFG = _load_admin_config()
FIXED_ADMIN_EMAIL = str(_ADMIN_CFG.get("email") or "admin@company.local").strip().lower()
FIXED_ADMIN_USERNAME = str(_ADMIN_CFG.get("username") or "admin").strip().lower()
FIXED_ADMIN_PASSWORD = str(_ADMIN_CFG.get("password") or "Admin@12345")
FIXED_ADMIN_DISPLAY_NAME = str(_ADMIN_CFG.get("display_name") or "System Admin").strip()


# In-memory stores for live progress and uploaded bytes.
_jobs: dict = {}
_prepared_pairs: dict = {}
_id_counter = [0]
_id_lock = threading.Lock()


def _session_now_ts() -> int:
    return int(datetime.now().timestamp())


def _start_user_session(user_id: int) -> None:
    session.clear()
    session.permanent = True
    now_ts = _session_now_ts()
    session["user_id"] = int(user_id)
    session["login_at"] = now_ts
    session["last_seen"] = now_ts


def _expire_session(reason: str = "session_expired") -> None:
    user_id = session.get("user_id")
    try:
        if user_id:
            db.add_audit_log(int(user_id), reason, {})
    except Exception:
        pass
    session.clear()


def _current_user_id():
    return session.get("user_id")


def _current_user():
    user_id = _current_user_id()
    if not user_id:
        return None
    return db.get_user_by_id(int(user_id))


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not _current_user_id():
            return jsonify({"error": "Authentication required"}), 401
        return fn(*args, **kwargs)
    return wrapper


def _is_admin(user: dict | None = None) -> bool:
    user = user or _current_user()
    return bool(user and (user.get("role") == "admin" or int(user.get("is_admin") or 0) == 1))


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = _current_user()
        if not user:
            return jsonify({"error": "Authentication required"}), 401
        if not _is_admin(user):
            return jsonify({"error": "Admin access required"}), 403
        return fn(*args, **kwargs)
    return wrapper


def _run_for_current_user(run_id: str):
    user = _current_user()
    if not user:
        return None
    if _is_admin(user):
        return db.get_run_by_id(run_id)
    return db.get_run_for_user(run_id, int(user["id"]))


def _file_pair_for_current_user(file_pair_id: int):
    user = _current_user()
    if not user:
        return None
    if _is_admin(user):
        return db.get_file_pair_by_id(file_pair_id)
    return db.get_file_pair_for_user(file_pair_id, int(user["id"]))


@app.before_request
def enforce_session_timeout():
    """Expire logged-in sessions after inactivity.

    Closing the browser tab does not reliably delete cookies, so the backend
    enforces timeout on every request. /api/me is allowed to return the new
    logged-out state instead of a hard 401.
    """
    # Static files and public auth routes must stay accessible.
    public_paths = {"/", "/advanced", "/simple", "/api/me", "/api/login", "/api/setup-admin", "/api/invite-info", "/api/accept-invite", "/api/forgot-password", "/api/reset-password-info", "/api/reset-password"}
    if request.path.startswith("/static/") or request.path in public_paths or request.path.startswith("/accept-invite/") or request.path.startswith("/reset-password/"):
        # Still expire if an old session exists, so /api/me reflects reality.
        pass

    user_id = session.get("user_id")
    if not user_id:
        return None

    now_ts = _session_now_ts()
    last_seen = int(session.get("last_seen", now_ts))

    if now_ts - last_seen > SESSION_IDLE_TIMEOUT_SECONDS:
        _expire_session("session_idle_timeout")
        if request.path == "/api/me":
            return None
        return jsonify({"error": "Session expired. Please log in again."}), 401

    session["last_seen"] = now_ts
    session.permanent = True
    return None


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _safe_text(value) -> str:
    value = "" if value is None else str(value)
    value = value.replace("\r\n", "\n").replace("\r", "\n").strip()
    if value.startswith(("=", "+", "-", "@")):
        value = "'" + value
    return value


def _next_qa_id():
    with _id_lock:
        _id_counter[0] += 1
        return f"QA-{_id_counter[0]:04d}"


def _normalize_pair_key(name_or_path, validation_mode="pdf") -> str:
    """Build a stable pair key using language/category/document name."""
    p = Path(str(name_or_path).replace("\\", "/"))
    stem = unicodedata.normalize("NFC", p.stem).casefold()
    language_names = {"english","spanish","french","russian","polish","vietnamese","chinese","arabic","farsi","urdu","german","portuguese","italian","dutch"}
    parts = [seg.casefold() for seg in p.parts]
    language = next((x for x in parts if x in language_names), "")
    category = ""
    if len(p.parts) >= 2:
        parent = p.parts[-2].casefold()
        if len(parent) <= 3 and parent.isalnum():
            category = parent
    # RTF packages need language + category + basename because multiple localized
    # documents can share a title. PDF/HTML and PDF/DOCX packages use basename
    # pairing because the PDF side may be flat while HTML carries language folders.
    if validation_mode == "rtf":
        key = "/".join(x for x in (language, category, stem) if x)
    else:
        key = stem
    return "".join(ch for ch in key if ch.isalnum())


def _sanitize_name(name: str, max_len: int = 40) -> str:
    stem = Path(name or "Uploaded_Batch").stem
    stem = unicodedata.normalize("NFKD", stem)
    stem = "".join(ch for ch in stem if not unicodedata.combining(ch))
    stem = stem.strip().replace(" ", "_")
    stem = re.sub(r"[^A-Za-z0-9_\-]+", "", stem)
    stem = re.sub(r"_+", "_", stem).strip("_-")
    if not stem:
        stem = "Uploaded_Batch"
    return stem[:max_len]


def _next_batch_run_number() -> int:
    max_num = 0
    if BATCH_RUNS.exists():
        for folder in BATCH_RUNS.iterdir():
            if not folder.is_dir():
                continue
            m = re.match(r"BATCH_(\d{3})_", folder.name)
            if m:
                max_num = max(max_num, int(m.group(1)))
    return max_num + 1


def _create_batch_run_folder(source_zip_name: str) -> dict:
    with _id_lock:
        batch_num = _next_batch_run_number()
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        safe_source_name = _sanitize_name(source_zip_name, max_len=40)
        folder_name = f"BATCH_{batch_num:03d}_{timestamp}_{safe_source_name}"

        batch_dir = BATCH_RUNS / folder_name
        reports_dir = batch_dir / "reports"
        audits_dir = batch_dir / "audits"
        metadata_dir = batch_dir / "metadata"

        reports_dir.mkdir(parents=True, exist_ok=True)
        audits_dir.mkdir(parents=True, exist_ok=True)
        metadata_dir.mkdir(parents=True, exist_ok=True)

    return {
        "run_id": folder_name,
        "batch_dir": str(batch_dir),
        "reports_dir": str(reports_dir),
        "audits_dir": str(audits_dir),
        "metadata_dir": str(metadata_dir),
    }


def _infer_source_zip_name(files, items=None, source_type="pdf") -> str:
    names = [f.filename for f in files if f and f.filename]
    source_type = (source_type or "pdf").lower()
    source_ext = f".{source_type}"

    for name in names:
        if Path(name).suffix.lower() == ".zip" and source_type in Path(name).stem.lower():
            return name
    for name in names:
        if Path(name).suffix.lower() == ".zip":
            return name
    if items:
        for item in items:
            if Path(item.get("filename", "")).suffix.lower() == source_ext:
                return item.get("filename", "")
    return names[0] if names else "Uploaded_Batch"


def _infer_html_zip_name(files) -> str:
    names = [f.filename for f in files if f and f.filename]
    for name in names:
        if Path(name).suffix.lower() == ".zip" and "html" in Path(name).stem.lower():
            return name
    return ""


def _safe_relpath(name: str, fallback: str) -> Path:
    name = (name or fallback or "uploaded_file").replace("\\", "/").lstrip("/")
    parts = []
    for part in name.split("/"):
        part = part.strip()
        if not part or part in {".", ".."}:
            continue
        if part.startswith("__MACOSX") or part.startswith("."):
            continue
        parts.append(part)
    if not parts:
        parts = [Path(fallback or "uploaded_file").name]
    return Path(*parts)


def _validate_upload_extension(filename: str) -> None:
    ext = Path(filename or "").suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTS:
        raise ValueError(f"Unsupported file type: {filename}")


def _read_uploaded_file_items(file_storage, relpath: str):
    """Read uploads and recursively inspect nested ZIP packages in memory."""
    original_name = file_storage.filename or "uploaded_file"
    _validate_upload_extension(original_name)
    uploaded_bytes = file_storage.read()

    def read_zip(blob, depth=0):
        if depth > 4:
            raise ValueError("ZIP nesting is too deep.")
        items=[]; total_uncompressed=0
        try:
            with zipfile.ZipFile(io.BytesIO(blob)) as z:
                infos=[i for i in z.infolist() if not i.is_dir()]
                if len(infos)>MAX_ZIP_FILES: raise ValueError(f"ZIP has too many files: {original_name}")
                for info in infos:
                    name=info.filename.replace("\\","/")
                    if name.startswith("__MACOSX/") or Path(name).name.startswith("."): continue
                    total_uncompressed += int(info.file_size or 0)
                    if total_uncompressed>MAX_ZIP_UNCOMPRESSED_BYTES: raise ValueError(f"ZIP is too large after extraction: {original_name}")
                    data=z.read(info)
                    if Path(name).suffix.lower()==".zip":
                        items.extend(read_zip(data, depth+1))
                    else:
                        safe=_safe_relpath(name, Path(name).name)
                        ext=safe.suffix.lower()
                        # Package-level files that must never become QA candidates.
                        low=str(safe).replace("\\","/").lower()
                        if safe.name.lower()=="epicdesktopindex.html" or "epicdesktop" in low or safe.name.lower()=="mytonomyepicmetadata.xml":
                            continue
                        if ext in {".xml",".css",".js",".json"}: continue
                        _validate_upload_extension(safe.name)
                        items.append({"filename": safe.name, "relpath": safe.as_posix(), "data": data})
        except zipfile.BadZipFile:
            raise ValueError(f"Invalid ZIP file: {original_name}")
        return items

    if Path(original_name).suffix.lower()==".zip":
        return read_zip(uploaded_bytes)
    safe=_safe_relpath(relpath, original_name)
    _validate_upload_extension(safe.name)
    if safe.name.lower()=="epicdesktopindex.html" or "epicdesktop" in safe.as_posix().lower():
        return []
    return [{"filename": safe.name, "relpath": safe.as_posix(), "data": uploaded_bytes}]


def _write_prepared_pair_to_temp(prepared: dict, tmpdir: str):
    """Materialize one source/HTML pair into a temp folder for qa_engine, then delete after job."""
    tmp = Path(tmpdir)
    source_item = prepared["source"]
    html_item = prepared["html"]
    assets = prepared.get("assets", [])

    source_path = tmp / Path(source_item["filename"]).name
    html_path = tmp / Path(html_item["filename"]).name
    source_path.write_bytes(source_item["data"])
    html_path.write_bytes(html_item["data"])

    images_dir = tmp / "Images"
    images_dir.mkdir(exist_ok=True)
    asset_by_basename = {}

    for asset in assets:
        asset_name = Path(asset["filename"]).name
        asset_path = images_dir / asset_name
        if not asset_path.exists():
            asset_path.write_bytes(asset["data"])
        asset_by_basename.setdefault(asset_name.lower(), asset_path)

    # Rewrite HTML image src by basename to temp Images folder.
    try:
        from bs4 import BeautifulSoup
        raw = html_path.read_text(encoding="utf-8", errors="replace")
        soup = BeautifulSoup(raw, "html.parser")
        changed = False
        for img in soup.find_all("img"):
            src = img.get("src", "")
            basename = Path(src).name.lower()
            if not basename:
                continue
            asset_path = asset_by_basename.get(basename)
            if not asset_path:
                continue
            new_src = os.path.relpath(asset_path, html_path.parent).replace("\\", "/")
            if img.get("src") != new_src:
                img["src"] = new_src
                changed = True
        if changed:
            html_path.write_text(str(soup), encoding="utf-8")
    except Exception as e:
        print(f"Warning: could not rewrite temp HTML image paths: {e}")

    return str(source_path), str(html_path)


def _append_processing_log(audits_dir: Path, message: str) -> None:
    audits_dir = Path(audits_dir)
    audits_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with (audits_dir / "processing_log.txt").open("a", encoding="utf-8") as f:
        f.write(f"[{ts}] {message}\n")


def _write_json(path: Path, data: dict) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


CRITICAL_CATEGORIES = {
    "Changed Text", "Missing Text", "Extra Text", "Title Mismatch", "Missing Title",
    "Broken Image", "Missing Image", "Image Mismatch", "Image Content Mismatch",
    "Garbled Text", "Encoding / Mojibake", "Unexpected Closing Tag", "Unclosed Tag",
    "HTML Parse Error", "Missing Punctuation", "Extra Punctuation", "Punctuation Mismatch",
}

MAJOR_CATEGORIES = {
    "Missing Heading Tag", "Heading Level Mismatch", "Missing B Tag", "Missing Required B Tag",
    "Missing Strong/Bold Tag", "Unexpected Strong/Bold Tag", "Missing Emphasis/Italic Tag",
    "Unexpected Emphasis/Italic Tag", "Bullet Marker Mismatch", "Missing Bullet Marker",
    "Missing List Item Tag", "Style Attribute Mismatch", "Tag Attribute Mismatch",
    "Text Outside Expected HTML Tags", "Broken Link", "Missing Language Attribute",
    "Missing RTL Direction", "RTL Direction Mismatch",
}

MINOR_CATEGORIES = {"Filename Mismatch", "Placeholder Text", "Placeholder", "Image Order Mismatch"}

CONTENT_CATEGORIES = {
    "Changed Text", "Missing Text", "Extra Text", "Missing Punctuation", "Extra Punctuation",
    "Punctuation Mismatch", "Garbled Text", "Encoding / Mojibake",
}
STRUCTURE_CATEGORIES = {
    "Missing Heading Tag", "Heading Level Mismatch", "Missing B Tag", "Missing Required B Tag",
    "Missing Strong/Bold Tag", "Unexpected Strong/Bold Tag", "Missing Emphasis/Italic Tag",
    "Unexpected Emphasis/Italic Tag", "Bullet Marker Mismatch", "Missing Bullet Marker",
    "Missing List Item Tag", "Tag Attribute Mismatch", "Unexpected Closing Tag", "Unclosed Tag",
    "HTML Parse Error",
}
STYLE_CATEGORIES = {"Style Attribute Mismatch", "Missing RTL Direction", "RTL Direction Mismatch"}
IMAGE_CATEGORIES = {"Broken Image", "Missing Image", "Image Mismatch", "Image Content Mismatch", "Image Order Mismatch"}
METADATA_CATEGORIES = {"Filename Mismatch", "Title Mismatch", "Missing Title", "Broken Link", "Missing Language Attribute"}


def _display_severity(issue) -> str:
    category = getattr(issue, "category", "") or ""
    raw = (getattr(issue, "severity", "") or "").lower()
    if category in CRITICAL_CATEGORIES:
        return "Critical"
    if category in MAJOR_CATEGORIES:
        return "Major"
    if category in MINOR_CATEGORIES:
        return "Minor"
    if raw == "error":
        return "Critical"
    if raw == "warning":
        return "Major"
    return "Minor"


def _issue_area(issue) -> str:
    category = getattr(issue, "category", "") or ""
    if category in CONTENT_CATEGORIES:
        return "Content"
    if category in STRUCTURE_CATEGORIES:
        return "Structure"
    if category in STYLE_CATEGORIES:
        return "Style"
    if category in IMAGE_CATEGORIES:
        return "Image"
    if category in METADATA_CATEGORIES:
        return "Metadata"
    return "Other"


def _issue_to_dict(issue) -> dict:
    return {
        "category": getattr(issue, "category", "") or "",
        "severity": _display_severity(issue),
        "engine_severity": getattr(issue, "severity", "") or "",
        "area": _issue_area(issue),
        "line": getattr(issue, "line", None),
        "message": getattr(issue, "message", "") or "",
        "snippet": getattr(issue, "snippet", "") or "",
        "expected": getattr(issue, "expected", "") or "",
        "actual": getattr(issue, "actual", "") or "",
    }


def _counts_by_severity(issue_dicts: list[dict]) -> dict:
    return {
        "critical": sum(1 for i in issue_dicts if i.get("severity") == "Critical"),
        "major": sum(1 for i in issue_dicts if i.get("severity") == "Major"),
        "minor": sum(1 for i in issue_dicts if i.get("severity") == "Minor"),
    }


def _file_overall_severity(counts: dict) -> str:
    if counts.get("critical", 0) >= 1 or counts.get("major", 0) >= 5:
        return "Critical"
    if counts.get("major", 0) >= 1 or counts.get("minor", 0) >= 8:
        return "Major"
    if counts.get("minor", 0) >= 1:
        return "Minor"
    return "Passed"


def _summary_from_result(result: dict, pair: dict, uid: str, report_path: str, audit_path: str) -> dict:
    issue_dicts = [_issue_to_dict(i) for i in result.get("issues", [])]
    counts = _counts_by_severity(issue_dicts)
    total = len(issue_dicts)
    status = "Passed" if total == 0 else "Failed"
    file_severity = _file_overall_severity(counts)

    type_counts = {}
    area_counts = {}
    for issue in issue_dicts:
        type_counts[issue["category"]] = type_counts.get(issue["category"], 0) + 1
        area_counts[issue["area"]] = area_counts.get(issue["area"], 0) + 1

    return {
        "unique_id": uid,
        "run_id": pair["run_id"],
        "file_pair_db_id": pair["file_pair_db_id"],
        "pair_id": pair["pair_id"],
        "file_no": pair["file_no"],
        "source_type": pair.get("source_type", "pdf"),
        "source_name": pair.get("source_name", ""),
        "pdf_name": pair.get("source_name", ""),
        "html_name": pair.get("html_name", ""),
        "target_name": pair.get("target_name") or pair.get("html_name", ""),
        "validation_mode": pair.get("validation_mode") or pair.get("source_type", ""),
        "language": result.get("language", ""),
        "status": status,
        "file_severity": file_severity,
        "issue_count": total,
        "critical_count": counts["critical"],
        "major_count": counts["major"],
        "minor_count": counts["minor"],
        "type_counts": type_counts,
        "area_counts": area_counts,
        "report_path": report_path,
        "audit_path": audit_path,
        "issues": issue_dicts,
        "ran_by": pair.get("ran_by") or pair.get("username") or "",
    }


def _result_from_file_pair_row(fp: dict) -> dict:
    issues = db.list_issues_for_file_pair(int(fp["id"]))
    issue_dicts = [
        {
            "category": i.get("category") or "",
            "severity": i.get("severity") or "",
            "engine_severity": i.get("engine_severity") or "",
            "area": i.get("area") or "",
            "line": i.get("html_line") or "",
            "message": i.get("message") or "",
            "expected": i.get("expected") or "",
            "actual": i.get("actual") or "",
            "snippet": i.get("snippet") or "",
        }
        for i in issues
    ]
    return {
        "file_pair_db_id": fp["id"],
        "pair_id": fp.get("pair_id"),
        "run_id": fp.get("run_id"),
        "file_no": fp.get("file_no"),
        "source_type": fp.get("source_type"),
        "source_name": fp.get("source_name"),
        "pdf_name": fp.get("source_name"),
        "html_name": fp.get("html_name"),
        "target_name": fp.get("html_name"),
        "job_id": fp.get("job_id") or "",
        "language": fp.get("language") or "",
        "status": "Passed" if int(fp.get("issue_count") or 0) == 0 and fp.get("status") == "done" else ("Failed" if fp.get("status") == "done" else fp.get("status")),
        "file_severity": fp.get("file_severity") or "Passed",
        "issue_count": int(fp.get("issue_count") or 0),
        "critical_count": int(fp.get("critical_count") or 0),
        "major_count": int(fp.get("major_count") or 0),
        "minor_count": int(fp.get("minor_count") or 0),
        "report_path": fp.get("report_path") or "",
        "audit_path": fp.get("audit_path") or "",
        "issues": issue_dicts,
    }


def _generate_qa_workbook(output_path: str, run: dict, results: list[dict]) -> str:
    """Create the 3-sheet Excel report for any validation mode."""
    wb = Workbook()
    mode = str(run.get("source_type") or "pdf").lower()
    mode_label = {"pdf": "PDF ↔ HTML", "rtf": "RTF ↔ HTML", "pdf_word": "PDF ↔ WORD"}.get(mode, mode.upper())
    ws_summary = wb.active
    ws_summary.title = "Overall Summary"
    ws_files = wb.create_sheet("File Based Summary")
    ws_errors = wb.create_sheet("Individual File Errors")

    colors = {
        "navy": "0F172A", "white": "FFFFFF", "muted": "64748B",
        "light_bg": "F8FAFC", "section_bg": "EAF1FF", "section_fg": "1E3A8A",
        "header_bg": "1E293B", "header_fg": "FFFFFF", "border": "D0D7E2",
        "Critical_bg": "FEE2E2", "Critical_fg": "DC2626", "Critical_band": "FCA5A5",
        "Major_bg": "FFEDD5", "Major_fg": "C2410C", "Major_band": "FDBA74",
        "Minor_bg": "FEF9C3", "Minor_fg": "A16207", "Minor_band": "FDE68A",
        "Passed_bg": "DCFCE7", "Passed_fg": "15803D", "Passed_band": "86EFAC",
        "label_bg": "F1F5F9", "expected_bg": "EFF6FF", "actual_bg": "FFF7ED", "context_bg": "F8FAFC",
    }
    thin = Side(border_style="thin", color=colors["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, bg=None, fg="000000", bold=False, size=10, align="left", vertical="top", wrap=True, italic=False):
        cell.font = Font(name="Aptos", size=size, bold=bold, italic=italic, color=fg)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical=vertical, wrap_text=wrap)
        cell.border = border
        return cell

    def write_cell(ws, row, col, value="", bg=None, fg="000000", bold=False, size=10, align="left", vertical="top", wrap=True):
        return style_cell(ws.cell(row=row, column=col, value=_safe_text(value)), bg, fg, bold, size, align, vertical, wrap)

    def merge_row(ws, row, c1, c2, text, bg, fg="000000", bold=True, size=11, height=None, align="left"):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        style_cell(ws.cell(row=row, column=c1, value=_safe_text(text)), bg, fg, bold, size, align, "center", True)
        for c in range(c1 + 1, c2 + 1):
            style_cell(ws.cell(row=row, column=c), bg, fg, bold, size)
        if height:
            ws.row_dimensions[row].height = height

    def merge_detail_row(ws, row, label, value, bg, max_col):
        write_cell(ws, row, 1, label, bg=bg, fg="334155", bold=True, size=9, align="right", wrap=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max_col)
        style_cell(ws.cell(row=row, column=2, value=_safe_text(value) or "—"), bg, "0F172A", False, 9, "left", "top", True)
        for c in range(3, max_col + 1):
            style_cell(ws.cell(row=row, column=c), bg, "0F172A", False, 9)
        length = len(_safe_text(value))
        ws.row_dimensions[row].height = 90 if length > 450 else 60 if length > 220 else 34

    def sev_style(sev):
        sev = sev or "Minor"
        if sev == "Critical":
            return {"bg": colors["Critical_bg"], "fg": colors["Critical_fg"], "band": colors["Critical_band"]}
        if sev == "Major":
            return {"bg": colors["Major_bg"], "fg": colors["Major_fg"], "band": colors["Major_band"]}
        if sev == "Passed":
            return {"bg": colors["Passed_bg"], "fg": colors["Passed_fg"], "band": colors["Passed_band"]}
        return {"bg": colors["Minor_bg"], "fg": colors["Minor_fg"], "band": colors["Minor_band"]}

    def issue_sort_key(issue):
        order = {"Critical": 0, "Major": 1, "Minor": 2}
        return (order.get(issue.get("severity"), 9), issue.get("line") or 999999, issue.get("category") or "")

    for ws in [ws_summary, ws_files, ws_errors]:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = None

    results = sorted(results, key=lambda x: x.get("file_no") or 99999)
    total_files = len(results)
    passed_files = sum(1 for r in results if int(r.get("issue_count") or 0) == 0 and r.get("status") in {"Passed", "done"})
    failed_files = sum(1 for r in results if int(r.get("issue_count") or 0) > 0 or r.get("status") == "Failed")
    total_issues = sum(int(r.get("issue_count") or 0) for r in results)
    total_critical = sum(int(r.get("critical_count") or 0) for r in results)
    total_major = sum(int(r.get("major_count") or 0) for r in results)
    total_minor = sum(int(r.get("minor_count") or 0) for r in results)

    category_counter = {}
    file_sev_counter = {"Critical": 0, "Major": 0, "Minor": 0, "Passed": 0}
    for r in results:
        file_sev_counter[r.get("file_severity") or "Passed"] = file_sev_counter.get(r.get("file_severity") or "Passed", 0) + 1
        for issue in r.get("issues") or []:
            cat = issue.get("category") or "Unknown"
            category_counter[cat] = category_counter.get(cat, 0) + 1

    # Sheet 1
    for col, width in {1: 24, 2: 16, 3: 24, 4: 16, 5: 24, 6: 16}.items():
        ws_summary.column_dimensions[get_column_letter(col)].width = width
    row = 1
    merge_row(ws_summary, row, 1, 6, f"Document QA Validation Report · {mode_label} · {run.get('run_id')}", colors["navy"], colors["white"], True, 15, 32)
    row += 1
    merge_row(ws_summary, row, 1, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}    |    Overall Summary", "EEF2FF", colors["muted"], False, 10, 24)
    row += 2
    merge_row(ws_summary, row, 1, 6, "BATCH OVERVIEW", colors["section_bg"], colors["section_fg"], True, 11, 24)
    row += 1
    overview = [
        ("Source Type", str(run.get("source_type", "")).upper()),
        ("Total Files", total_files),
        ("Passed Files", passed_files),
        ("Failed Files", failed_files),
        ("Total Errors", total_issues),
        ("Run ID", run.get("run_id", "")),
        ("Source ZIP", run.get("source_zip_name", "")),
        ("Target Package", run.get("html_zip_name", "") or "—"),
    ]
    for idx, (label, value) in enumerate(overview):
        col = 1 if idx % 2 == 0 else 4
        write_cell(ws_summary, row, col, label, bg=colors["label_bg"], fg=colors["muted"], bold=True, align="center")
        ws_summary.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + 2)
        style_cell(ws_summary.cell(row=row, column=col + 1, value=_safe_text(value)), colors["white"], "000000", True, 10, "center", "center")
        for c in range(col + 2, col + 3):
            style_cell(ws_summary.cell(row=row, column=c), colors["white"])
        if idx % 2 == 1:
            row += 1
    row += 2
    merge_row(ws_summary, row, 1, 6, "ERROR SEVERITY DISTRIBUTION", colors["section_bg"], colors["section_fg"], True, 11, 24)
    row += 1
    for label, val, sev in [("Critical Errors", total_critical, "Critical"), ("Major Errors", total_major, "Major"), ("Minor Errors", total_minor, "Minor")]:
        s = sev_style(sev)
        write_cell(ws_summary, row, 1, label, bg=s["bg"], fg=s["fg"], bold=True)
        write_cell(ws_summary, row, 2, val, bg=s["bg"], fg=s["fg"], bold=True, align="center")
        row += 1
    row += 1
    merge_row(ws_summary, row, 1, 6, "FILE SEVERITY DISTRIBUTION", colors["section_bg"], colors["section_fg"], True, 11, 24)
    row += 1
    for sev in ["Critical", "Major", "Minor", "Passed"]:
        s = sev_style(sev)
        write_cell(ws_summary, row, 1, sev, bg=s["bg"], fg=s["fg"], bold=True)
        write_cell(ws_summary, row, 2, file_sev_counter.get(sev, 0), bg=s["bg"], fg=s["fg"], bold=True, align="center")
        row += 1
    row += 1
    merge_row(ws_summary, row, 1, 6, "ERROR CATEGORY BREAKDOWN", colors["section_bg"], colors["section_fg"], True, 11, 24)
    row += 1
    write_cell(ws_summary, row, 1, "Error Type", bg=colors["header_bg"], fg=colors["header_fg"], bold=True, align="center")
    write_cell(ws_summary, row, 2, "Count", bg=colors["header_bg"], fg=colors["header_fg"], bold=True, align="center")
    row += 1
    if category_counter:
        for cat, count in sorted(category_counter.items(), key=lambda x: (-x[1], x[0])):
            write_cell(ws_summary, row, 1, cat, bg=colors["white"])
            write_cell(ws_summary, row, 2, count, bg=colors["white"], bold=True, align="center")
            row += 1
    else:
        merge_row(ws_summary, row, 1, 6, "No errors found.", colors["Passed_bg"], colors["Passed_fg"], True, 10, 26)

    # Sheet 3 first for links.
    for col, width in {1: 8, 2: 28, 3: 14, 4: 14, 5: 12, 6: 20, 7: 20, 8: 20, 9: 20, 10: 20, 11: 20}.items():
        ws_errors.column_dimensions[get_column_letter(col)].width = width
    max_col = 11
    detail_row_by_file_no = {}
    row = 1
    merge_row(ws_errors, row, 1, max_col, f"Individual File Errors · {run.get('run_id')}", colors["navy"], colors["white"], True, 15, 32)
    row += 1
    merge_row(ws_errors, row, 1, max_col, "Each file is shown separately. Every error has Expected, Actual, and Context rows.", "EEF2FF", colors["muted"], False, 10, 28)
    row += 3
    if not results:
        merge_row(ws_errors, row, 1, max_col, "No completed QA results found.", colors["light_bg"], colors["muted"], True, 11, 26)
    for result in results:
        file_no = result.get("file_no") or "—"
        detail_row_by_file_no[file_no] = row
        sev = result.get("file_severity") or "Passed"
        s = sev_style(sev)
        source_name = result.get("source_name") or result.get("pdf_name") or "—"
        target_name = result.get("target_name") or result.get("html_name") or "—"
        merge_row(ws_errors, row, 1, max_col, f"FILE {file_no} — {source_name}  →  {target_name}", s["band"], s["fg"], True, 11, 32)
        row += 1
        meta = f"Status: {result.get('status') or '—'}    |    Severity: {sev}    |    Total: {result.get('issue_count', 0)}    |    Critical: {result.get('critical_count', 0)}    Major: {result.get('major_count', 0)}    Minor: {result.get('minor_count', 0)}    |    Language: {result.get('language') or '—'}"
        merge_row(ws_errors, row, 1, max_col, meta, s["bg"], s["fg"], True, 9, 26)
        row += 1
        merge_detail_row(ws_errors, row, "Source", source_name, colors["light_bg"], max_col)
        row += 1
        merge_detail_row(ws_errors, row, "Target", target_name, colors["light_bg"], max_col)
        row += 2
        issues = sorted(result.get("issues") or [], key=issue_sort_key)
        if not issues:
            merge_row(ws_errors, row, 1, max_col, "No issues found for this file.", colors["Passed_bg"], colors["Passed_fg"], True, 10, 28)
            row += 3
            continue
        for idx, issue in enumerate(issues, 1):
            issue_sev = issue.get("severity") or "Minor"
            istyle = sev_style(issue_sev)
            merge_row(ws_errors, row, 1, max_col, f"Error {idx}: {issue.get('category') or '—'} · {issue_sev}", istyle["bg"], istyle["fg"], True, 10, 24)
            row += 1
            headers = ["#", "Error Type", "Severity", "Area", "Target Location", "Description"]
            for c, h in enumerate(headers, 1):
                write_cell(ws_errors, row, c, h, bg=colors["header_bg"], fg=colors["header_fg"], bold=True, size=9, align="center", vertical="center")
            ws_errors.merge_cells(start_row=row, start_column=6, end_row=row, end_column=max_col)
            for c in range(7, max_col + 1):
                style_cell(ws_errors.cell(row=row, column=c), colors["header_bg"], colors["header_fg"], True, 9)
            row += 1
            write_cell(ws_errors, row, 1, idx, bg=colors["white"], bold=True, align="center", vertical="center")
            write_cell(ws_errors, row, 2, issue.get("category") or "—", bg=istyle["bg"], fg=istyle["fg"], bold=True, vertical="center")
            write_cell(ws_errors, row, 3, issue_sev, bg=istyle["bg"], fg=istyle["fg"], bold=True, align="center", vertical="center")
            write_cell(ws_errors, row, 4, issue.get("area") or "—", bg=colors["white"], align="center", vertical="center")
            write_cell(ws_errors, row, 5, issue.get("line") or "—", bg=colors["white"], align="center", vertical="center")
            ws_errors.merge_cells(start_row=row, start_column=6, end_row=row, end_column=max_col)
            style_cell(ws_errors.cell(row=row, column=6, value=_safe_text(issue.get("message") or "—")[:300]), colors["white"], "000000", False, 9, "left", "top", True)
            for c in range(7, max_col + 1):
                style_cell(ws_errors.cell(row=row, column=c), colors["white"])
            ws_errors.row_dimensions[row].height = 44
            row += 1
            merge_detail_row(ws_errors, row, "Expected", issue.get("expected") or "—", colors["expected_bg"], max_col)
            row += 1
            merge_detail_row(ws_errors, row, "Actual", issue.get("actual") or "—", colors["actual_bg"], max_col)
            row += 1
            merge_detail_row(ws_errors, row, "Context", issue.get("snippet") or issue.get("message") or "—", colors["context_bg"], max_col)
            row += 2
        row += 2

    # Sheet 2
    for col, width in {1: 8, 2: 44, 3: 16, 4: 16, 5: 12, 6: 12, 7: 12, 8: 12, 9: 16, 10: 18}.items():
        ws_files.column_dimensions[get_column_letter(col)].width = width
    row = 1
    merge_row(ws_files, row, 1, 10, f"File Based Summary · {run.get('run_id')}", colors["navy"], colors["white"], True, 15, 32)
    row += 1
    merge_row(ws_files, row, 1, 10, "One row per file. Use Details Link to jump to the file's error section.", "EEF2FF", colors["muted"], False, 10, 24)
    row += 2
    header_row = row
    headers = ["No", "File Pair", "Status", "File Severity", "Errors", "Critical", "Major", "Minor", "Language", "Details Link"]
    for c, h in enumerate(headers, 1):
        write_cell(ws_files, row, c, h, bg=colors["header_bg"], fg=colors["header_fg"], bold=True, size=9, align="center", vertical="center")
    row += 1
    if not results:
        merge_row(ws_files, row, 1, 10, "No completed QA results found.", colors["light_bg"], colors["muted"], True, 10, 26)
    for result in results:
        file_no = result.get("file_no") or "—"
        sev = result.get("file_severity") or "Passed"
        s = sev_style(sev)
        file_pair = f"{result.get('source_name') or result.get('pdf_name') or ''}\n→ {result.get('target_name') or result.get('html_name') or ''}"
        values = [file_no, file_pair, result.get("status") or "—", sev, result.get("issue_count", 0), result.get("critical_count", 0), result.get("major_count", 0), result.get("minor_count", 0), result.get("language") or "—", "Open Details"]
        for c, value in enumerate(values, 1):
            bg = s["bg"] if c in (3, 4) else colors["white"]
            fg = s["fg"] if c in (3, 4) else "000000"
            cell = write_cell(ws_files, row, c, value, bg=bg, fg=fg, bold=c in (1, 3, 4, 5, 6, 7, 8, 10), size=9, align="center" if c != 2 else "left", vertical="center")
            if c == 10 and detail_row_by_file_no.get(file_no):
                cell.hyperlink = f"#'Individual File Errors'!A{detail_row_by_file_no[file_no]}"
                cell.font = Font(name="Aptos", size=9, bold=True, color="2563EB", underline="single")
        ws_files.row_dimensions[row].height = 42
        row += 1
    if row - 1 >= header_row:
        ws_files.auto_filter.ref = f"A{header_row}:J{row - 1}"

    for ws in [ws_summary, ws_files, ws_errors]:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.sheet_view.zoomScale = 90

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _write_file_pairs_csv(metadata_dir: Path, pairs: list[dict]) -> None:
    import csv
    with (Path(metadata_dir) / "file_pairs.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["file_no", "pair_id", "source_file", "html_file", "status"])
        for idx, p in enumerate(pairs, 1):
            writer.writerow([idx, p.get("pair_id", ""), p.get("source_name") or p.get("pdf_name", ""), p.get("target_name") or p.get("html_name", ""), "matched"])


def _write_unmatched_csv(metadata_dir: Path, unmatched: list) -> None:
    import csv
    with (Path(metadata_dir) / "unmatched_files.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["type", "side", "filename", "reason"])
        for item in unmatched:
            if isinstance(item, dict):
                writer.writerow(["unmatched", item.get("side", ""), item.get("filename", ""), item.get("reason", "No corresponding file found")])
            else:
                writer.writerow(["unmatched", "", item, "No corresponding file found"])


def _write_batch_manifest(run: dict) -> None:
    run_row = db.get_run_by_id(run["run_id"]) or run
    file_pairs = db.list_file_pairs_for_run(run["run_id"])
    _write_json(Path(run["metadata_dir"]) / "batch_manifest.json", {"run": run_row, "file_pairs": file_pairs, "generated_at": _now()})


@app.route("/")
def index():
    return send_from_directory(str(STATIC), "index.html")


@app.route("/advanced")
def advanced_page():
    return send_from_directory(str(STATIC), "index.html")


@app.route("/simple")
def simple_page():
    return send_from_directory(str(STATIC), "index.html")


@app.route("/accept-invite/<token>")
def accept_invite_page(token):
    return send_from_directory(str(STATIC), "index.html")


@app.route("/reset-password/<token>")
def reset_password_page(token):
    return send_from_directory(str(STATIC), "index.html")


@app.route("/api/me")
def api_me():
    user = _current_user()
    return jsonify({
        "authenticated": bool(user),
        "user": user,
        "has_users": db.has_users(),
    })


def _password_strength_error(password: str, username: str = "") -> str | None:
    username = (username or "").lower().strip()
    password = password or ""
    if len(password) < 10:
        return "Password must be at least 10 characters."
    if len(password) > 128:
        return "Password is too long. Use 128 characters or fewer."
    if any(ch.isspace() for ch in password):
        return "Password must not contain spaces."
    if not re.search(r"[a-z]", password):
        return "Password must include at least one lowercase letter."
    if not re.search(r"[A-Z]", password):
        return "Password must include at least one uppercase letter."
    if not re.search(r"\d", password):
        return "Password must include at least one number."
    if not re.search(r"[^A-Za-z0-9]", password):
        return "Password must include at least one symbol, such as !, @, #, or %."
    if username and username in password.lower():
        return "Password must not contain your username."
    if password.lower() in {"password", "password123", "admin123", "qwerty123", "letmein", "welcome123", "12345678"}:
        return "Password is too common. Choose a stronger password."
    return None


def _valid_email(email: str) -> bool:
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", (email or "").strip().lower()))


def _valid_username(username: str) -> bool:
    return bool(re.fullmatch(r"[a-zA-Z0-9_.-]{3,40}", (username or "").strip()))


@app.route("/api/setup-admin", methods=["POST"])
def api_setup_admin():
    return jsonify({
        "error": "Admin setup page is disabled. The single admin account is seeded from server configuration."
    }), 403


@app.route("/api/login", methods=["POST"])

def api_login():
    data = request.json or {}
    login = (data.get("login") or data.get("username") or data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    login_mode = (data.get("login_mode") or "user").strip().lower()

    user = db.get_user_by_login(login)
    valid = bool(user and check_password_hash(user["password_hash"], password))

    if valid:
        is_admin_login = user.get("role") == "admin" or int(user.get("is_admin") or 0) == 1
        if login_mode == "admin" and not is_admin_login:
            valid = False
        if login_mode == "user" and is_admin_login:
            valid = False

    if not valid:
        db.add_audit_log(None, "login_failed", {"login": login, "login_mode": login_mode}, ip_address=request.remote_addr or "")
        return jsonify({"error": "Invalid login or password."}), 401

    _start_user_session(user["id"])
    db.update_last_login(user["id"])
    db.add_audit_log(user["id"], "login_success", {"login": login, "login_mode": login_mode}, ip_address=request.remote_addr or "")
    return jsonify({"ok": True, "user": db.get_user_by_id(user["id"])})


@app.route("/api/logout", methods=["POST"])
@login_required
def api_logout():
    user_id = int(session["user_id"])
    db.add_audit_log(user_id, "logout", {}, ip_address=request.remote_addr or "")
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/invite-info")
def api_invite_info():
    token = request.args.get("token") or ""
    invite = db.get_invite_by_token(token)
    if not invite:
        return jsonify({"error": "Invalid invite link."}), 404
    return jsonify({
        "email": invite.get("email"),
        "role": invite.get("role"),
        "valid": bool(invite.get("valid")),
        "invalid_reason": invite.get("invalid_reason", ""),
        "expires_at": invite.get("expires_at"),
    })


@app.route("/api/accept-invite", methods=["POST"])
def api_accept_invite():
    data = request.json or {}
    token = data.get("token") or ""
    username = (data.get("username") or "").strip().lower()
    password = data.get("password") or ""
    invite = db.get_invite_by_token(token)
    if not invite or not invite.get("valid"):
        return jsonify({"error": (invite or {}).get("invalid_reason") or "Invalid invite link."}), 400
    if not _valid_username(username):
        return jsonify({"error": "Username must be 3-40 characters and use letters, numbers, dot, dash, or underscore."}), 400
    strength_error = _password_strength_error(password, username)
    if strength_error:
        return jsonify({"error": strength_error}), 400
    try:
        user_id = db.consume_invite(token, username=username, password_hash=generate_password_hash(password))
        db.add_audit_log(user_id, "invite_accepted", {"email": invite.get("email"), "username": username}, ip_address=request.remote_addr or "")
    except Exception as e:
        return jsonify({"error": f"Could not create account: {e}"}), 400
    _start_user_session(user_id)
    return jsonify({"ok": True, "user": db.get_user_by_id(user_id)})


@app.route("/api/forgot-password", methods=["POST"])
def api_forgot_password():
    """Public, self-service reset REQUEST only.

    Deliberately never reveals whether the account exists and never returns
    a usable token to the caller -- an admin must generate and share the
    actual reset link, the same trust model this portal already uses for
    invites. Always responds with the same generic message so this endpoint
    can't be used to enumerate valid usernames/emails.
    """
    data = request.json or {}
    login = (data.get("login") or data.get("username") or data.get("email") or "").strip().lower()
    generic_message = "If that account exists, your administrator has been notified and will send you a reset link."
    if login:
        try:
            req = db.create_password_reset_request(login, request_ip=request.remote_addr or "")
            if req:
                db.add_audit_log(int(req["user_id"]), "password_reset_requested", {"login": login}, "password_reset", str(req["id"]), request.remote_addr or "")
        except Exception:
            pass
    return jsonify({"ok": True, "message": generic_message})


@app.route("/api/reset-password-info")
def api_reset_password_info():
    token = request.args.get("token") or ""
    pr = db.get_password_reset_by_token(token)
    if not pr:
        return jsonify({"error": "Invalid reset link."}), 404
    return jsonify({
        "username": pr.get("username_snapshot"),
        "valid": bool(pr.get("valid")),
        "invalid_reason": pr.get("invalid_reason", ""),
        "expires_at": pr.get("expires_at"),
    })


@app.route("/api/reset-password", methods=["POST"])
def api_reset_password():
    data = request.json or {}
    token = data.get("token") or ""
    password = data.get("password") or ""
    pr = db.get_password_reset_by_token(token)
    if not pr or not pr.get("valid"):
        return jsonify({"error": (pr or {}).get("invalid_reason") or "Invalid reset link."}), 400
    strength_error = _password_strength_error(password, pr.get("username_snapshot") or "")
    if strength_error:
        return jsonify({"error": strength_error}), 400
    try:
        user_id = db.consume_password_reset(token, generate_password_hash(password))
        db.add_audit_log(user_id, "password_reset_completed", {}, "password_reset", str(pr["id"]), request.remote_addr or "")
    except Exception as e:
        return jsonify({"error": f"Could not reset password: {e}"}), 400
    return jsonify({"ok": True})


@app.route("/api/admin/summary")
@admin_required
def api_admin_summary():
    return jsonify(db.admin_summary())


@app.route("/api/admin/users")
@admin_required
def api_admin_users():
    return jsonify({"users": db.list_users()})


@app.route("/api/admin/users/<int:user_id>/status", methods=["POST"])
@admin_required
def api_admin_user_status(user_id):
    data = request.json or {}
    status = data.get("status") or "disabled"
    target = db.get_any_user_by_id(user_id)
    if target and (target.get("role") == "admin" or int(target.get("is_admin") or 0) == 1):
        return jsonify({"error": "The fixed admin account cannot be disabled from the portal."}), 400
    db.set_user_status(user_id, status)
    db.add_audit_log(int(session["user_id"]), "user_status_changed", {"target_user_id": user_id, "status": status}, "user", str(user_id), request.remote_addr or "")
    return jsonify({"ok": True})


@app.route("/api/admin/users/<int:user_id>/role", methods=["POST"])

@admin_required
def api_admin_user_role(user_id):
    return jsonify({"error": "Role changes are disabled. This portal supports exactly one fixed admin; invited accounts are normal users."}), 403


@app.route("/api/admin/invites")

@admin_required
def api_admin_invites():
    return jsonify({"invites": db.list_invites()})


@app.route("/api/admin/invites", methods=["POST"])
@admin_required
def api_admin_create_invite():
    data = request.json or {}
    email = (data.get("email") or "").strip().lower()
    role = "user"
    if not _valid_email(email):
        return jsonify({"error": "Enter a valid company email."}), 400
    try:
        invite = db.create_invite(email=email, role="user", created_by_admin_id=int(session["user_id"]))
        invite["invite_url"] = request.host_url.rstrip("/") + "/accept-invite/" + invite["token"]
        db.add_audit_log(int(session["user_id"]), "invite_created", {"email": email, "role": "user"}, "invite", str(invite["id"]), request.remote_addr or "")
        return jsonify({"ok": True, "invite": invite})
    except Exception as e:
        return jsonify({"error": f"Could not create invite: {e}"}), 400


@app.route("/api/admin/invites/<int:invite_id>/cancel", methods=["POST"])
@admin_required
def api_admin_cancel_invite(invite_id):
    db.cancel_invite(invite_id)
    db.add_audit_log(int(session["user_id"]), "invite_cancelled", {}, "invite", str(invite_id), request.remote_addr or "")
    return jsonify({"ok": True})


@app.route("/api/admin/password-resets")
@admin_required
def api_admin_password_resets():
    status = request.args.get("status", "requested")
    if status == "all":
        status = None
    return jsonify({"requests": db.list_password_reset_requests(status=status, limit=200)})


@app.route("/api/admin/password-resets/<int:request_id>/generate-link", methods=["POST"])
@admin_required
def api_admin_generate_reset_link(request_id):
    try:
        result = db.generate_password_reset_link(request_id, admin_id=int(session["user_id"]))
        result["reset_url"] = request.host_url.rstrip("/") + "/reset-password/" + result["token"]
        db.add_audit_log(int(session["user_id"]), "password_reset_link_generated", {}, "password_reset", str(request_id), request.remote_addr or "")
        return jsonify({"ok": True, "reset": result})
    except Exception as e:
        return jsonify({"error": f"Could not generate reset link: {e}"}), 400


@app.route("/api/admin/password-resets/<int:request_id>/cancel", methods=["POST"])
@admin_required
def api_admin_cancel_reset(request_id):
    db.cancel_password_reset_request(request_id)
    db.add_audit_log(int(session["user_id"]), "password_reset_cancelled", {}, "password_reset", str(request_id), request.remote_addr or "")
    return jsonify({"ok": True})


@app.route("/api/admin/runs")
@admin_required
def api_admin_runs():
    return jsonify({"runs": db.list_all_runs(limit=300)})


@app.route("/api/admin/problem-runs")
@admin_required
def api_admin_problem_runs():
    return jsonify({"runs": db.list_problem_runs(limit=300)})


@app.route("/api/admin/audit-logs")
@admin_required
def api_admin_audit_logs():
    return jsonify({"logs": db.list_audit_logs(limit=300)})

@app.route("/upload", methods=["POST"])
@login_required
def upload():
    user=_current_user(); user_id=int(user["id"]); session_id=str(uuid.uuid4())[:8]
    upload_instance_id=(request.form.get("upload_instance_id") or str(uuid.uuid4())).strip()
    mode=(request.form.get("source_type") or "pdf").lower().strip()
    if mode not in ALLOWED_SOURCE_TYPES:
        return jsonify({"error":"Unsupported validation mode. Use pdf, rtf, or pdf_word."}),400
    files=request.files.getlist("files"); relpaths=request.form.getlist("relpaths")
    if len(relpaths)!=len(files): relpaths=[f.filename for f in files]
    items=[]
    try:
        for f,relpath in zip(files,relpaths):
            if f and f.filename: items.extend(_read_uploaded_file_items(f,relpath))
    except ValueError as e:
        return jsonify({"error":str(e)}),400

    sources={}; targets={}; assets=[]
    if mode=="pdf": source_exts={".pdf"}; target_exts={".html",".htm"}
    elif mode=="rtf": source_exts={".rtf"}; target_exts={".html",".htm"}
    else: source_exts={".pdf",".docx"}; target_exts={".pdf",".docx"}

    for item in items:
        ext=Path(item["filename"]).suffix.lower(); low=item["relpath"].replace("\\","/").lower()
        if "epicdesktop" in low or Path(item["filename"]).name.lower()=="epicdesktopindex.html": continue
        if mode in {"pdf","rtf"} and ext in source_exts:
            sources.setdefault(_normalize_pair_key(item["relpath"],mode),[]).append(item)
        elif mode in {"pdf","rtf"} and ext in target_exts:
            targets.setdefault(_normalize_pair_key(item["relpath"],mode),[]).append(item)
        elif mode=="pdf_word" and ext in source_exts:
            sources.setdefault(_normalize_pair_key(item["relpath"],mode),[]).append(item)
        elif mode=="pdf_word" and ext in target_exts:
            targets.setdefault(_normalize_pair_key(item["relpath"],mode),[]).append(item)
        elif ext in ALLOWED_IMAGE_EXTS:
            assets.append(item)

    # PDF↔Word: one key can contain both formats; split deterministically by extension.
    if mode=="pdf_word":
        grouped={}
        for item in items:
            ext=Path(item["filename"]).suffix.lower()
            if ext in {".pdf",".docx"}:
                grouped.setdefault(_normalize_pair_key(item["relpath"],mode),[]).append(item)
        sources={}; targets={}
        for key,vals in grouped.items():
            pdfs=[v for v in vals if Path(v["filename"]).suffix.lower()==".pdf"]
            docs=[v for v in vals if Path(v["filename"]).suffix.lower()==".docx"]
            if pdfs and docs:
                sources[key]=pdfs; targets[key]=docs

    pairs=[]; unmatched_details=[]
    for key in sorted(set(sources)&set(targets)):
        sl=sources[key]; tl=targets[key]; count=min(len(sl),len(tl))
        for idx in range(count):
            src,tgt=sl[idx],tl[idx]; pair_id=f"{session_id}_{key or 'pair'}_{idx}"
            prepared={"source":src,"html":tgt,"target":tgt,"assets":assets,"source_name":src["filename"],"pdf_name":src["filename"],"target_name":tgt["filename"],"html_name":tgt["filename"],"source_type":mode,"validation_mode":mode,"session":session_id}
            _prepared_pairs[pair_id]=prepared
            pairs.append({"pair_id":pair_id,"source_name":src["filename"],"pdf_name":src["filename"],"target_name":tgt["filename"],"html_name":tgt["filename"],"source_type":mode,"validation_mode":mode,"source_path":f"memory://{pair_id}/source","target_path":f"memory://{pair_id}/target","html_path":f"memory://{pair_id}/target","image_count":len(assets)})
        for v in sl[count:]:
            unmatched_details.append({"side":"Source","filename":v["filename"],"reason":"No corresponding target file found"})
        for v in tl[count:]:
            unmatched_details.append({"side":"Target","filename":v["filename"],"reason":"No corresponding source file found"})
    for key in sorted(set(sources)-set(targets)):
        for v in sources[key]: unmatched_details.append({"side":"Source","filename":v["filename"],"reason":"No corresponding target file found"})
    for key in sorted(set(targets)-set(sources)):
        for v in targets[key]: unmatched_details.append({"side":"Target","filename":v["filename"],"reason":"No corresponding source file found"})
    unmatched=[x["filename"] for x in unmatched_details]

    source_zip_name=_infer_source_zip_name(files,items,source_type=mode)
    html_zip_name=_infer_html_zip_name(files)
    batch_ctx=_create_batch_run_folder(source_zip_name); run_id=batch_ctx["run_id"]
    batch_report_path=str(Path(batch_ctx["reports_dir"])/"batch_report.xlsx")
    run_record={"run_id":run_id,"user_id":user_id,"username_snapshot":user.get("username",""),"email_snapshot":user.get("email",""),"source_type":mode,"source_zip_name":source_zip_name,"html_zip_name":html_zip_name,"batch_dir":batch_ctx["batch_dir"],"reports_dir":batch_ctx["reports_dir"],"audits_dir":batch_ctx["audits_dir"],"metadata_dir":batch_ctx["metadata_dir"],"batch_report_path":batch_report_path,"total_files":len(pairs),"prepared_files":len(pairs),"status":"prepared","status_message":"Files uploaded and paired. Waiting for QA run.","created_at":_now()}
    db.create_run(run_record); pairs=db.insert_file_pairs(run_id,user_id,mode,pairs)
    for p in pairs:
        if p["pair_id"] in _prepared_pairs: _prepared_pairs[p["pair_id"]].update({"run_id":run_id,"file_no":p["file_no"],"file_pair_db_id":p["file_pair_db_id"]})
    _write_file_pairs_csv(Path(batch_ctx["metadata_dir"]),pairs); _write_unmatched_csv(Path(batch_ctx["metadata_dir"]),unmatched); _write_batch_manifest(run_record)
    _append_processing_log(Path(batch_ctx["audits_dir"]),f"Batch created by {user['username']}: {run_id}")
    _append_processing_log(Path(batch_ctx["audits_dir"]),f"Prepared {len(pairs)} pair(s).")
    if unmatched: _append_processing_log(Path(batch_ctx["audits_dir"]),f"Found {len(unmatched)} unmatched file(s).")
    db.add_audit_log(user_id,"batch_uploaded",{"run_id":run_id,"upload_instance_id":upload_instance_id,"pairs":len(pairs),"unmatched":len(unmatched),"mode":mode,"created_at":run_record["created_at"]})
    # Never reuse a previous QA result: each /upload creates a fresh run and fresh file-pair rows.
    return jsonify({"pairs":pairs,"unmatched":unmatched,"unmatched_details":unmatched_details,"session":session_id,"batch_id":run_id,"run_id":run_id,"source_type":mode,"validation_mode":mode,"created_at":run_record["created_at"],"fresh_run":True})


@app.route("/run", methods=["POST"])

@login_required
def run():
    user_id = int(session["user_id"])
    runner_user = db.get_user_by_id(user_id) or {"username": ""}
    runner_username = runner_user.get("username", "")
    data = request.json or {}
    pair_id = data.get("pair_id", "")
    run_id = data.get("batch_id") or data.get("run_id") or ""

    run_row = db.get_run_for_user(run_id, user_id)
    if not run_row:
        return jsonify({"error": "Run not found or access denied."}), 404

    prepared = _prepared_pairs.get(pair_id)
    if not prepared:
        return jsonify({"error": "Prepared file pair not found. Upload again if the server was restarted."}), 400

    file_pair = db.get_file_pair_by_pair_id(pair_id)
    if not file_pair or int(file_pair["user_id"]) != user_id:
        return jsonify({"error": "File pair not found or access denied."}), 404

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {"status": "running", "progress": 0, "label": "Starting…", "result": None, "error": None, "user_id": user_id, "run_id": run_id, "file_pair_db_id": file_pair["id"]}
    db.update_file_pair_running(int(file_pair["id"]), job_id)
    db.update_run_status(run_id, "running")

    def worker():
        try:
            def prog(pct, label):
                _jobs[job_id]["progress"] = pct
                _jobs[job_id]["label"] = label

            with tempfile.TemporaryDirectory() as tmpdir:
                source_path, target_path = _write_prepared_pair_to_temp(prepared, tmpdir)
                result = run_checks(
                    source_path,
                    target_path,
                    progress_cb=prog,
                    validation_mode=prepared.get("validation_mode") or run_row.get("source_type") or "pdf",
                )

                uid = _next_qa_id()
                file_no = int(file_pair["file_no"])
                reports_dir = Path(run_row["reports_dir"])
                audits_dir = Path(run_row["audits_dir"])
                report_path = reports_dir / f"file_{file_no:03d}_report.xlsx"
                audit_path = audits_dir / f"file_{file_no:03d}_audit.json"

                pair_context = {
                    "run_id": run_id,
                    "file_pair_db_id": int(file_pair["id"]),
                    "pair_id": pair_id,
                    "file_no": file_no,
                    "source_type": prepared.get("source_type", run_row["source_type"]),
                    "source_name": prepared.get("source_name") or file_pair["source_name"],
                    "html_name": prepared.get("html_name") or file_pair["html_name"],
                    "target_name": prepared.get("target_name") or prepared.get("html_name") or file_pair["html_name"],
                    "validation_mode": prepared.get("validation_mode", run_row.get("source_type")),
                    "ran_by": runner_username,
                }
                summary_result = _summary_from_result(result, pair_context, uid, str(report_path), str(audit_path))

                _write_json(audit_path, {"validator_version": VALIDATOR_VERSION, "generated_at": _now(), "result": summary_result})
                db.replace_issues_for_file_pair(int(file_pair["id"]), run_id, summary_result["issues"])
                db.update_file_pair_done(int(file_pair["id"]), summary_result)
                db.update_run_counts(run_id)

                # Every per-file report is also a 3-sheet workbook, just scoped to one file.
                _generate_qa_workbook(str(report_path), run_row, [summary_result])
                _append_processing_log(audits_dir, f"Completed file {file_no}: {pair_context['source_name']} -> {pair_context.get('target_name') or pair_context.get('html_name', '')}")
                _write_batch_manifest(dict(run_row))

            _jobs[job_id]["status"] = "done"
            _jobs[job_id]["progress"] = 100
            _jobs[job_id]["label"] = "Completed"
            _jobs[job_id]["result"] = summary_result
            db.add_audit_log(user_id, "qa_completed", {"run_id": run_id, "file_pair_id": file_pair["id"], "issues": summary_result["issue_count"]})
        except Exception as e:
            import traceback
            err = str(e) + "\n\n" + traceback.format_exc()
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = err
            db.update_file_pair_error(int(file_pair["id"]), err)
            db.update_run_counts(run_id)
            try:
                _append_processing_log(Path(run_row["audits_dir"]), f"ERROR job {job_id}: {e}")
                db.add_audit_log(user_id, "qa_error", {"run_id": run_id, "file_pair_id": file_pair["id"], "error": str(e)})
            except Exception:
                pass

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/status/<job_id>")
@login_required
def status(job_id):
    job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    if int(job.get("user_id", -1)) != int(session["user_id"]):
        return jsonify({"error": "Access denied"}), 403
    return jsonify(job)


@app.route("/batch-summary/<run_id>")
@login_required
def batch_summary_route(run_id):
    run_row = _run_for_current_user(run_id)
    if not run_row:
        return jsonify({"error": "Unknown run or access denied"}), 404
    files = db.list_file_pairs_for_run(run_id)
    results = [_result_from_file_pair_row(fp) for fp in files if fp.get("status") == "done"]
    return jsonify({"batch_id": run_id, "run": run_row, "results": results, "files": files})


@app.route("/download/<job_id>")
@login_required
def download(job_id):
    job = _jobs.get(job_id)
    if not job or job.get("status") != "done":
        return "Not ready", 404
    if int(job.get("user_id", -1)) != int(session["user_id"]):
        return "Access denied", 403
    path = Path((job.get("result") or {}).get("report_path", ""))
    if not path.exists():
        return f"Report file not found: {path}", 404
    return send_file(str(path), as_attachment=True, download_name=path.name)


@app.route("/download-file/<int:file_pair_id>")
@login_required
def download_file(file_pair_id):
    fp = _file_pair_for_current_user(file_pair_id)
    if not fp:
        return "File not found or access denied", 404
    user_id = int(session["user_id"])
    path = Path(fp.get("report_path") or "")
    if not path.exists():
        # Regenerate from database if path is missing.
        run_row = _run_for_current_user(fp["run_id"])
        result = _result_from_file_pair_row(fp)
        path = Path(run_row["reports_dir"]) / f"file_{int(fp['file_no']):03d}_report.xlsx"
        _generate_qa_workbook(str(path), run_row, [result])
    return send_file(str(path), as_attachment=True, download_name=path.name)


@app.route("/download-batch/<run_id>")
@login_required
def download_batch(run_id):
    user_id = int(session["user_id"])
    run_row = _run_for_current_user(run_id)
    if not run_row:
        return "Run not found or access denied", 404

    file_pairs = db.list_file_pairs_for_run(run_id)
    results = [_result_from_file_pair_row(fp) for fp in file_pairs if fp.get("status") == "done"]
    output_path = Path(run_row["reports_dir"]) / "batch_report.xlsx"
    _generate_qa_workbook(str(output_path), run_row, results)
    db.set_batch_report_path(run_id, str(output_path))
    _append_processing_log(Path(run_row["audits_dir"]), "Generated batch Excel report.")
    db.add_audit_log(user_id, "batch_report_downloaded", {"run_id": run_id})
    return send_file(str(output_path), as_attachment=True, download_name=output_path.name)


@app.route("/api/history")
@login_required
def api_history():
    runs = db.list_runs_for_user(int(session["user_id"]), limit=100)
    return jsonify({"runs": runs})


@app.route("/api/history/<run_id>")
@login_required
def api_history_detail(run_id):
    run_row = _run_for_current_user(run_id)
    if not run_row:
        return jsonify({"error": "Run not found or access denied"}), 404
    files = db.list_file_pairs_for_run(run_id)
    return jsonify({"run": run_row, "files": files})


@app.route("/api/file-report/<int:file_pair_id>")
@login_required
def api_file_report(file_pair_id):
    """
    Returns the saved issue list for a single file pair.

    The frontend renders issues straight from the QA-run response in the
    normal flow; this route is the fallback path for older run payloads
    that only carried summary counts (issue_count/critical_count/etc.)
    without the underlying issue rows, so "Open Report" never shows
    "No issues found" for a run that actually recorded errors.
    """
    fp = _file_pair_for_current_user(file_pair_id)
    if not fp:
        return jsonify({"error": "File pair not found or access denied"}), 404
    return jsonify(_result_from_file_pair_row(fp))


@app.route("/run-batch", methods=["POST"])
def run_batch():
    return "Plain form batch upload is disabled. Use the main UI.", 410


@app.route("/download-direct/<filename>")
def download_direct(filename):
    return "Legacy direct downloads are disabled. Use /download/<job_id>, /download-file/<file_pair_id>, or /download-batch/<run_id>.", 410


if __name__ == "__main__":
    db.init_db()
    db.ensure_fixed_admin(
        email=FIXED_ADMIN_EMAIL,
        username=FIXED_ADMIN_USERNAME,
        password_hash=generate_password_hash(FIXED_ADMIN_PASSWORD),
        display_name=FIXED_ADMIN_DISPLAY_NAME,
    )
    port = int(os.environ.get("QA_TOOL_PORT", "7331"))
    if os.environ.get("QA_TOOL_NO_BROWSER", "0") != "1":
        threading.Timer(1.3, lambda: webbrowser.open(f"http://localhost:{port}")).start()
    print(f"\n{'=' * 58}")
    print(f"  Source → HTML Batch QA Tool  —  http://localhost:{port}")
    print(f"  SQLite DB: {db.DB_PATH}")
    print(f"  Admin config: {ADMIN_CONFIG_PATH}")
    print(f"  Fixed admin login: {FIXED_ADMIN_EMAIL} or {FIXED_ADMIN_USERNAME}")
    if FIXED_ADMIN_PASSWORD == "Admin@12345":
        print("  Default admin password: Admin@12345  (change QA_ADMIN_PASSWORD for real use)")
    print(f"{'=' * 58}\n")
    app.run(port=port, debug=False, threaded=True)